import openpyxl


class HomePageData:

    @staticmethod
    def get_test_data():
        book = openpyxl.load_workbook("C:\\Users\\7J6295897\\PycharmProjects\\PythonSelFramework\\testdata"
                                      "\\pyexcellDemo.xlsx")
        sheet = book.active
        data = []
        for i in range(2, sheet.max_row + 1):
            dictionary = {}
            for j in range(2, sheet.max_column + 1):
                dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data.append(dictionary)
        return data
