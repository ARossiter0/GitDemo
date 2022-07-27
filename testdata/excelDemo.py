import openpyxl

book = openpyxl.load_workbook("pyexcellDemo.xlsx")

sheet = book.active

data = []

cell = sheet.cell(row=1, column=2)
print(cell.value)

# sheet.cell(row=2, column=2).value = "Alex"
# print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)

for i in range(2, sheet.max_row+1):
    dictionary = {}
    for j in range(2, sheet.max_column+1):
        dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    data.append(dictionary)

print(data)
